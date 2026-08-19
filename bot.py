# ============================================================
# GGOD FATHERR VCF MAKER BOT  ⚡  PREMIUM EDITION
# Admin: @ggod_fatherr
# ============================================================

import os
import io
import csv
import logging
import asyncio
import openpyxl
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)

# ── Logging ───────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────
BOT_TOKEN      = os.environ.get("BOT_TOKEN", "8603833578:AAFT3eB6HiDMyq8t_litDCw3Z2rFzB4LD4Q")
PREMIUM_CODE   = os.environ.get("PREMIUM_CODE", "#000#")
ADMIN_USERNAME = "@ggod_fatherr"
ADMIN_URL      = "https://t.me/ggod_fatherr"

# ── In-memory stores ──────────────────────────────────────────
premium_users: set = set()

# ── Branding ──────────────────────────────────────────────────
FOOTER  = "\n\n⚡ Powered by @ggod_fatherr"
DIVIDER = "▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰▰"
THIN    = "─────────────────────"

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def is_premium(user_id: int) -> bool:
    return user_id in premium_users

def file_stem(fname: str) -> str:
    if "." in fname:
        return fname.rsplit(".", 1)[0]
    return fname or "output"

def premium_required_msg() -> str:
    return (
        "🔒 Premium Access Required\n"
        + THIN + "\n"
        "This feature is locked for free users.\n\n"
        "💎 Get Premium: " + ADMIN_USERNAME
        + FOOTER
    )

# ─────────────────────────────────────────────────────────────
# ANIMATED LOADING HELPER
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
# KEYBOARDS
# ─────────────────────────────────────────────────────────────

def main_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📁 TXT → VCF",       callback_data="cmd_txt_to_vcf"),
            InlineKeyboardButton("📊 TXT/VCF → CSV",   callback_data="cmd_txtvcf_to_csv"),
        ],
        [
            InlineKeyboardButton("📇 CSV → VCF",       callback_data="cmd_csv_to_vcf"),
            InlineKeyboardButton("📄 VCF → TXT",       callback_data="cmd_vcf_to_txt"),
        ],
        [
            InlineKeyboardButton("💬 MSG → TXT",       callback_data="cmd_msg_to_txt"),
            InlineKeyboardButton("📝 Rename File",     callback_data="cmd_rename_file"),
        ],
        [
            InlineKeyboardButton("✏️ Rename Contacts", callback_data="cmd_rename_ctc"),
            InlineKeyboardButton("🧩 Merge VCFs",      callback_data="cmd_merge_vcf"),
        ],
        [
            InlineKeyboardButton("📚 Merge TXTs",      callback_data="cmd_merge_txt"),
            InlineKeyboardButton("✂️ Split File",      callback_data="cmd_split_file"),
        ],
        [
            InlineKeyboardButton("🛡️ Navy Format",     callback_data="cmd_admin_navy"),
            InlineKeyboardButton("📱 Number Compiler", callback_data="cmd_num_compiler"),
        ],
        [
            InlineKeyboardButton("🔄 Restart",         callback_data="cmd_reset"),
            InlineKeyboardButton("ℹ️ Help & Guide",    callback_data="help"),
        ],
        [
            InlineKeyboardButton("👤 Contact Admin",   url=ADMIN_URL),
        ],
    ])

def shortcut_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("📁 TXT→VCF"),  KeyboardButton("📄 VCF→TXT"),  KeyboardButton("📇 CSV→VCF")],
            [KeyboardButton("📊 →CSV"),     KeyboardButton("📝 Rename"),    KeyboardButton("✏️ Ren.CTC")],
            [KeyboardButton("🧩 MergeVCF"), KeyboardButton("📚 MergeTXT"), KeyboardButton("✂️ Split")],
            [KeyboardButton("🛡️ Navy"),     KeyboardButton("ℹ️ Help"),      KeyboardButton("🔄 Reset")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Choose a feature or type a command..."
    )

def back_keyboard(feature_cb: str = None) -> InlineKeyboardMarkup:
    rows = []
    if feature_cb:
        rows.append([InlineKeyboardButton("◀️ Back", callback_data=feature_cb)])
    rows.append([InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_menu")])
    return InlineKeyboardMarkup(rows)

def output_fmt_keyboard(feature_cb: str = None) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton("📄 TXT",  callback_data="fmt_txt"),
            InlineKeyboardButton("📇 VCF",  callback_data="fmt_vcf"),
        ],
        [
            InlineKeyboardButton("📊 CSV",  callback_data="fmt_csv"),
            InlineKeyboardButton("📗 XLSX", callback_data="fmt_xlsx"),
        ],
    ]
    if feature_cb:
        rows.append([InlineKeyboardButton("◀️ Back", callback_data=feature_cb)])
    rows.append([InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_menu")])
    return InlineKeyboardMarkup(rows)

def done_keyboard(feature_cb: str = None) -> InlineKeyboardMarkup:
    rows = []
    if feature_cb:
        rows.append([InlineKeyboardButton("🔁 Use Again", callback_data=feature_cb)])
    rows.append([InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_menu")])
    return InlineKeyboardMarkup(rows)

def compiler_fmt_keyboard(feature_cb: str = None) -> InlineKeyboardMarkup:
    """Output format picker for Number Compiler — includes XLS option."""
    rows = [
        [
            InlineKeyboardButton("📄 TXT",  callback_data="cmp_txt"),
            InlineKeyboardButton("📇 VCF",  callback_data="cmp_vcf"),
            InlineKeyboardButton("📊 CSV",  callback_data="cmp_csv"),
        ],
        [
            InlineKeyboardButton("📗 XLSX", callback_data="cmp_xlsx"),
            InlineKeyboardButton("📘 XLS",  callback_data="cmp_xls"),
        ],
    ]
    if feature_cb:
        rows.append([InlineKeyboardButton("◀️ Back", callback_data=feature_cb)])
    rows.append([InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_menu")])
    return InlineKeyboardMarkup(rows)

# ─────────────────────────────────────────────────────────────
# MESSAGE TEMPLATES
# ─────────────────────────────────────────────────────────────

def welcome_text() -> str:
    return (
        "╔══════════════════════════════╗\n"
        "║   👑  GGOD FATHERR BOT  👑  ║\n"
        "║   ⚡  VCF MAKER PREMIUM  ⚡  ║\n"
        "╚══════════════════════════════╝\n\n"
        "🔐 Premium Features Locked\n"
        + THIN + "\n"
        "🌟 Convert, merge, split & rename\n"
        "   contact files with ease.\n\n"
        "💎 Buy Access: " + ADMIN_USERNAME + "\n"
        + THIN + "\n"
        "👇 Tap below to unlock your access"
        + FOOTER
    )

def dashboard_text() -> str:
    return (
        "╔══════════════════════════════╗\n"
        "║  ⚡  GGOD FATHERR PREMIUM  ⚡ ║\n"
        "║  ✅  ALL FEATURES UNLOCKED   ║\n"
        "╚══════════════════════════════╝\n\n"
        + DIVIDER + "\n"
        "🚀 What would you like to do?\n"
        + DIVIDER
        + FOOTER
    )

def help_text() -> str:
    return (
        "╔══════════════════════════════╗\n"
        "║       ℹ️   HELP  GUIDE       ║\n"
        "╚══════════════════════════════╝\n\n"
        + THIN + "\n"
        "📌 Available Commands\n"
        + THIN + "\n"
        "▸ /start — Launch the bot\n"
        "▸ /txt_to_vcf — Numbers → VCF\n"
        "▸ /txtvcf_to_csv — TXT/VCF → CSV\n"
        "▸ /csv_to_vcf — CSV → VCF\n"
        "▸ /vcf_to_txt — VCF → Numbers\n"
        "▸ /msg_to_txt — Message → TXT\n"
        "▸ /rename_file — Rename any file\n"
        "▸ /rename_ctc — Rename VCF contacts\n"
        "▸ /merge_vcf — Merge VCF files\n"
        "▸ /merge_txt — Merge TXT files\n"
        "▸ /split_file — Split contacts\n"
        "▸ /admin_navy_file — Navy format\n"
        "▸ /done — Finalize merge\n"
        "▸ /reset — Clear session\n"
        + THIN + "\n"
        "👤 Admin: " + ADMIN_USERNAME
        + FOOTER
    )

# ─────────────────────────────────────────────────────────────
# CONVERSION LOGIC
# ─────────────────────────────────────────────────────────────

def txt_to_vcf(text: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    vcf_entries = []
    for i, number in enumerate(lines, 1):
        vcf_entries.append(
            "BEGIN:VCARD\nVERSION:3.0\n"
            "FN:Contact " + str(i) + "\nTEL:" + number + "\nEND:VCARD"
        )
    return "\n".join(vcf_entries)

def vcf_to_txt(text: str) -> str:
    numbers = []
    for line in text.splitlines():
        line = line.strip()
        if line.upper().startswith("TEL"):
            parts = line.split(":")
            if len(parts) >= 2:
                numbers.append(parts[-1].strip())
    return "\n".join(numbers)

def parse_vcf(text: str) -> list:
    entries = []
    name, phone = "Unknown", ""
    for line in text.splitlines():
        line = line.strip()
        if line.upper().startswith("FN:"):
            name = line[3:].strip()
        elif line.upper().startswith("TEL"):
            parts = line.split(":")
            if len(parts) >= 2:
                phone = parts[-1].strip()
        elif line.upper() == "END:VCARD":
            if phone:
                entries.append((name, phone))
            name, phone = "Unknown", ""
    return entries

def csv_to_vcf(text: str) -> str:
    reader = csv.DictReader(io.StringIO(text))
    vcf_entries = []
    i = 1
    for row in reader:
        name  = (row.get("Name") or row.get("name") or
                 row.get("FN") or ("Contact " + str(i)))
        phone = (row.get("Phone") or row.get("phone") or
                 row.get("TEL") or row.get("Number") or "")
        if phone:
            vcf_entries.append(
                "BEGIN:VCARD\nVERSION:3.0\n"
                "FN:" + name + "\nTEL:" + phone + "\nEND:VCARD"
            )
        i += 1
    return "\n".join(vcf_entries)

def rename_vcf_contacts(text: str, prefix: str) -> str:
    lines = text.splitlines()
    result = []
    counter = 1
    for line in lines:
        if line.strip().upper().startswith("FN:"):
            result.append("FN:" + prefix + " " + str(counter))
            counter += 1
        else:
            result.append(line)
    return "\n".join(result)

def split_contacts(text: str, count: int, ext: str) -> list:
    chunks = []
    if ext == "vcf":
        blocks = []
        current = []
        for line in text.splitlines():
            current.append(line)
            if line.strip().upper() == "END:VCARD":
                blocks.append("\n".join(current))
                current = []
        for i in range(0, len(blocks), count):
            chunks.append("\n".join(blocks[i:i + count]))
    else:
        lines = [l for l in text.splitlines() if l.strip()]
        for i in range(0, len(lines), count):
            chunks.append("\n".join(lines[i:i + count]))
    return chunks

def admin_navy_format(text: str) -> str:
    entries = parse_vcf(text)
    vcf_entries = []
    for i, (name, phone) in enumerate(entries, 1):
        vcf_entries.append(
            "BEGIN:VCARD\nVERSION:3.0\n"
            "FN:NAVY " + str(i).zfill(4) + " " + name + "\n"
            "TEL:" + phone + "\nEND:VCARD"
        )
    return "\n".join(vcf_entries)

def content_to_csv_bytes(content: str, source_ext: str) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Phone"])
    if source_ext == "vcf":
        for name, phone in parse_vcf(content):
            writer.writerow([name, phone])
    else:
        for i, line in enumerate(
            [l.strip() for l in content.splitlines() if l.strip()], 1
        ):
            writer.writerow(["Contact " + str(i), line])
    return output.getvalue().encode("utf-8")

def content_to_xlsx_bytes(content: str, source_ext: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Name", "Phone"])
    if source_ext == "vcf":
        for name, phone in parse_vcf(content):
            ws.append([name, phone])
    else:
        for i, line in enumerate(
            [l.strip() for l in content.splitlines() if l.strip()], 1
        ):
            ws.append(["Contact " + str(i), line])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def convert_content(content: str, source_ext: str, target_ext: str) -> tuple:
    if target_ext == "txt":
        data = vcf_to_txt(content).encode("utf-8") if source_ext == "vcf" else content.encode("utf-8")
        return data, "text/plain"
    elif target_ext == "vcf":
        if source_ext == "txt":
            data = txt_to_vcf(content).encode("utf-8")
        elif source_ext in ("csv", "xlsx"):
            data = csv_to_vcf(content).encode("utf-8")
        else:
            data = content.encode("utf-8")
        return data, "text/vcard"
    elif target_ext == "csv":
        return content_to_csv_bytes(content, source_ext), "text/csv"
    elif target_ext == "xlsx":
        return content_to_xlsx_bytes(content, source_ext), \
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return content.encode("utf-8"), "application/octet-stream"

def count_contacts(content: str, norm_ext: str) -> int:
    """Count total contacts/numbers in any normalised content."""
    if norm_ext == "vcf":
        return content.upper().count("BEGIN:VCARD")
    elif norm_ext in ("csv", "xlsx"):
        # Count non-header rows that have a phone column
        rows = list(csv.reader(io.StringIO(content)))
        if len(rows) <= 1:
            return 0
        return len([r for r in rows[1:] if any(c.strip() for c in r)])
    else:
        # txt — count non-empty lines
        return len([l for l in content.splitlines() if l.strip()])

# ─────────────────────────────────────────────────────────────
# MOBILE NUMBER COMPILER LOGIC
# ─────────────────────────────────────────────────────────────

import re as _re

# ── Country code lookup table ─────────────────────────────────
# Maps dialing prefix (string of digits) -> (country_name, language_code)
# Ordered longest-first so +1868 matches before +1
_CC_TABLE = {
    # 3-digit codes
    "355": ("Albania", "sq"),       "213": ("Algeria", "ar"),
    "376": ("Andorra", "ca"),       "244": ("Angola", "pt"),
    "374": ("Armenia", "hy"),       "994": ("Azerbaijan", "az"),
    "973": ("Bahrain", "ar"),       "880": ("Bangladesh", "bn"),
    "375": ("Belarus", "be"),       "501": ("Belize", "en"),
    "229": ("Benin", "fr"),         "975": ("Bhutan", "dz"),
    "591": ("Bolivia", "es"),       "387": ("Bosnia", "bs"),
    "267": ("Botswana", "en"),      "673": ("Brunei", "ms"),
    "359": ("Bulgaria", "bg"),      "226": ("Burkina Faso", "fr"),
    "257": ("Burundi", "fr"),       "855": ("Cambodia", "km"),
    "237": ("Cameroon", "fr"),      "238": ("Cape Verde", "pt"),
    "236": ("Central African Republic", "fr"),
    "235": ("Chad", "fr"),          "593": ("Ecuador", "es"),
    "503": ("El Salvador", "es"),   "240": ("Equatorial Guinea", "es"),
    "291": ("Eritrea", "ti"),       "372": ("Estonia", "et"),
    "251": ("Ethiopia", "am"),      "679": ("Fiji", "en"),
    "358": ("Finland", "fi"),       "241": ("Gabon", "fr"),
    "220": ("Gambia", "en"),        "995": ("Georgia", "ka"),
    "233": ("Ghana", "en"),         "350": ("Gibraltar", "en"),
    "299": ("Greenland", "kl"),     "502": ("Guatemala", "es"),
    "224": ("Guinea", "fr"),        "245": ("Guinea-Bissau", "pt"),
    "592": ("Guyana", "en"),        "509": ("Haiti", "fr"),
    "504": ("Honduras", "es"),      "852": ("Hong Kong", "zh"),
    "354": ("Iceland", "is"),       "353": ("Ireland", "en"),
    "972": ("Israel", "he"),        "225": ("Ivory Coast", "fr"),
    "876": ("Jamaica", "en"),       "962": ("Jordan", "ar"),
    "254": ("Kenya", "sw"),         "686": ("Kiribati", "en"),
    "383": ("Kosovo", "sq"),        "965": ("Kuwait", "ar"),
    "996": ("Kyrgyzstan", "ky"),    "856": ("Laos", "lo"),
    "371": ("Latvia", "lv"),        "961": ("Lebanon", "ar"),
    "266": ("Lesotho", "st"),       "231": ("Liberia", "en"),
    "218": ("Libya", "ar"),         "423": ("Liechtenstein", "de"),
    "370": ("Lithuania", "lt"),     "352": ("Luxembourg", "lb"),
    "853": ("Macau", "zh"),         "389": ("North Macedonia", "mk"),
    "261": ("Madagascar", "mg"),    "265": ("Malawi", "en"),
    "960": ("Maldives", "dv"),      "223": ("Mali", "fr"),
    "356": ("Malta", "mt"),         "692": ("Marshall Islands", "en"),
    "222": ("Mauritania", "ar"),    "230": ("Mauritius", "en"),
    "373": ("Moldova", "ro"),       "976": ("Mongolia", "mn"),
    "382": ("Montenegro", "sr"),    "212": ("Morocco", "ar"),
    "258": ("Mozambique", "pt"),    "264": ("Namibia", "en"),
    "674": ("Nauru", "na"),         "977": ("Nepal", "ne"),
    "505": ("Nicaragua", "es"),     "227": ("Niger", "fr"),
    "234": ("Nigeria", "en"),       "850": ("North Korea", "ko"),
    "968": ("Oman", "ar"),          "680": ("Palau", "en"),
    "507": ("Panama", "es"),        "675": ("Papua New Guinea", "en"),
    "595": ("Paraguay", "es"),      "970": ("Palestine", "ar"),
    "351": ("Portugal", "pt"),      "974": ("Qatar", "ar"),
    "242": ("Congo", "fr"),         "243": ("DR Congo", "fr"),
    "250": ("Rwanda", "rw"),        "685": ("Samoa", "sm"),
    "378": ("San Marino", "it"),    "239": ("Sao Tome", "pt"),
    "966": ("Saudi Arabia", "ar"),  "221": ("Senegal", "fr"),
    "381": ("Serbia", "sr"),        "232": ("Sierra Leone", "en"),
    "386": ("Slovenia", "sl"),      "677": ("Solomon Islands", "en"),
    "252": ("Somalia", "so"),       "211": ("South Sudan", "en"),
    "249": ("Sudan", "ar"),         "597": ("Suriname", "nl"),
    "268": ("Eswatini", "ss"),      "46": ("Sweden", "sv"),
    "963": ("Syria", "ar"),         "992": ("Tajikistan", "tg"),
    "255": ("Tanzania", "sw"),      "228": ("Togo", "fr"),
    "676": ("Tonga", "to"),         "868": ("Trinidad and Tobago", "en"),
    "216": ("Tunisia", "ar"),       "993": ("Turkmenistan", "tk"),
    "688": ("Tuvalu", "en"),        "256": ("Uganda", "en"),
    "380": ("Ukraine", "uk"),       "971": ("UAE", "ar"),
    "598": ("Uruguay", "es"),       "998": ("Uzbekistan", "uz"),
    "678": ("Vanuatu", "bi"),       "379": ("Vatican", "it"),
    "967": ("Yemen", "ar"),         "260": ("Zambia", "en"),
    "263": ("Zimbabwe", "en"),
    # 2-digit codes
    "86": ("China", "zh"),          "91": ("India", "hi"),
    "92": ("Pakistan", "ur"),       "44": ("UK", "en"),
    "49": ("Germany", "de"),        "33": ("France", "fr"),
    "39": ("Italy", "it"),          "34": ("Spain", "es"),
    "81": ("Japan", "ja"),          "82": ("South Korea", "ko"),
    "55": ("Brazil", "pt"),         "52": ("Mexico", "es"),
    "61": ("Australia", "en"),      "64": ("New Zealand", "en"),
    "27": ("South Africa", "en"),   "20": ("Egypt", "ar"),
    "62": ("Indonesia", "id"),      "63": ("Philippines", "tl"),
    "84": ("Vietnam", "vi"),        "66": ("Thailand", "th"),
    "90": ("Turkey", "tr"),         "98": ("Iran", "fa"),
    "48": ("Poland", "pl"),         "31": ("Netherlands", "nl"),
    "32": ("Belgium", "nl"),        "41": ("Switzerland", "de"),
    "43": ("Austria", "de"),        "47": ("Norway", "no"),
    "45": ("Denmark", "da"),        "30": ("Greece", "el"),
    "36": ("Hungary", "hu"),        "38": ("Ukraine", "uk"),
    "40": ("Romania", "ro"),        "7":  ("Russia", "ru"),
    "60": ("Malaysia", "ms"),       "65": ("Singapore", "en"),
    "95": ("Myanmar", "my"),        "94": ("Sri Lanka", "si"),
    "93": ("Afghanistan", "ps"),    "96": ("Libya", "ar"),
    "97": ("Jordan", "ar"),         "54": ("Argentina", "es"),
    "56": ("Chile", "es"),          "57": ("Colombia", "es"),
    "51": ("Peru", "es"),           "58": ("Venezuela", "es"),
    "53": ("Cuba", "es"),           "1":  ("USA", "en"),
}

def detect_country(num: str) -> tuple:
    """
    Detect country from a phone number.
    Returns (country_code_digits, local_number, country_name, lang_code).
    e.g. "+923001234567" -> ("92", "3001234567", "Pakistan", "ur")
    """
    digits = _re.sub(r'\D', '', num)
    # Remove leading zeros
    digits = digits.lstrip('0') or digits

    # Try longest match first (3 digits, then 2, then 1)
    for prefix_len in (3, 2, 1):
        prefix = digits[:prefix_len]
        if prefix in _CC_TABLE:
            country_name, lang_code = _CC_TABLE[prefix]
            local_number = digits[prefix_len:]
            return prefix, local_number, country_name, lang_code

    # Fallback: first 2 digits as country code, unknown country
    cc = digits[:2] if len(digits) >= 2 else digits
    return cc, digits[2:], "Unknown", "en"


def is_valid_number(s: str) -> bool:
    """Accept only strings that look like phone numbers (digits, +, -, spaces, parens)."""
    s = s.strip()
    if not s:
        return False
    cleaned = _re.sub(r'[\s\+\-\(\)\.]', '', s)
    return cleaned.isdigit() and len(cleaned) >= 6


def normalize_number(s: str) -> str:
    """Remove spaces and formatting but keep leading + for country code."""
    s = s.strip()
    return _re.sub(r'[\s\-\(\)\.]', '', s)

def compile_numbers_to_xlsx(numbers: list, tag: str = "") -> bytes:
    """
    Build the WhatsApp-import-ready XLSX.
    Columns: First Name | Last Name | Mobile Number | Language Code | Country | Email | Groups
    - Mobile Number = local number (country code stripped)
    - Language Code = auto-detected from country
    - Country = auto-detected country name
    - Groups = indiana 2000
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ["First Name", "Last Name", "Mobile Number",
               "Language Code", "Country", "Email", "Groups"]
    ws.append(headers)

    for i, num in enumerate(numbers, 1):
        cc, local, country_name, lang_code = detect_country(num)
        ws.append([
            "ND" + str(i).zfill(4),   # First Name
            "",                         # Last Name
            local,                      # Mobile Number (local, no country code)
            lang_code,                  # Language Code (auto-detected)
            country_name,               # Country (auto-detected)
            "",                         # Email
            "indiana 2000",             # Groups
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def compile_numbers_to_output(numbers: list, fmt: str, tag: str = "") -> tuple:
    """
    Convert compiled numbers list to the requested output format.
    Columns: First Name | Last Name | Mobile Number | Language Code | Country | Email | Groups
    Returns (bytes, filename_ext).
    """
    if fmt in ("xlsx", "xls"):
        data = compile_numbers_to_xlsx(numbers, tag)
        return data, fmt

    elif fmt == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["First Name", "Last Name", "Mobile Number",
                    "Language Code", "Country", "Email", "Groups"])
        for i, num in enumerate(numbers, 1):
            cc, local, country_name, lang_code = detect_country(num)
            w.writerow(["ND" + str(i).zfill(4), "", local,
                        lang_code, country_name, "", "indiana 2000"])
        return out.getvalue().encode("utf-8"), "csv"

    elif fmt == "vcf":
        entries = []
        for i, num in enumerate(numbers, 1):
            cc, local, country_name, lang_code = detect_country(num)
            entries.append(
                "BEGIN:VCARD\nVERSION:3.0\n"
                "FN:ND" + str(i).zfill(4) + "\n"
                "TEL:+" + cc + local + "\n"
                "CATEGORIES:indiana 2000\nEND:VCARD"
            )
        return "\n".join(entries).encode("utf-8"), "vcf"

    else:  # txt
        return "\n".join(numbers).encode("utf-8"), "txt"

# ─────────────────────────────────────────────────────────────
# UNIVERSAL FILE READER  (.txt / .csv / .vcf / .xlsx / .xls)
# ─────────────────────────────────────────────────────────────

ALL_FORMATS     = ("txt", "csv", "vcf", "xlsx", "xls")
COMPILER_FMTS   = ("txt", "csv", "vcf", "xlsx", "xls")

def read_file_content(raw: bytes, ext: str) -> tuple:
    """
    Read raw file bytes and return (text_content, normalised_ext).
    XLSX/XLS is converted to CSV-style text so all downstream logic works.
    """
    if ext in ("xlsx", "xls"):
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        out = io.StringIO()
        writer = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(c) if c is not None else "" for c in row])
        return out.getvalue(), "csv"
    else:
        return raw.decode("utf-8", errors="ignore"), ext

# ─────────────────────────────────────────────────────────────
# FEATURE PROMPTS
# ─────────────────────────────────────────────────────────────

FEATURE_PROMPTS = {
    "txt_to_vcf":    ("cmd_txt_to_vcf",    "📁 Any → VCF\n\nSend a .txt / .csv / .vcf / .xlsx file.\nPhone numbers will be converted to VCF contacts."),
    "txtvcf_to_csv": ("cmd_txtvcf_to_csv", "📊 Any → CSV/XLSX\n\nSend a .txt / .csv / .vcf / .xlsx file to convert."),
    "csv_to_vcf":    ("cmd_csv_to_vcf",    "📇 Any → VCF\n\nSend a .txt / .csv / .vcf / .xlsx file to convert to VCF contacts."),
    "vcf_to_txt":    ("cmd_vcf_to_txt",    "📄 Any → TXT\n\nSend a .txt / .csv / .vcf / .xlsx file to extract phone numbers."),
    "msg_to_txt":    ("cmd_msg_to_txt",    "💬 MSG → TXT\n\nType or paste your message below."),
    "rename_file":   ("cmd_rename_file",   "📝 Rename File\n\nSend any file (.txt / .csv / .vcf / .xlsx) to rename."),
    "rename_ctc":    ("cmd_rename_ctc",    "✏️ Rename Contacts\n\nSend a .vcf / .csv / .xlsx file to rename all contacts."),
    "merge_vcf":     ("cmd_merge_vcf",     "🧩 Merge → VCF\n\nSend .txt / .csv / .vcf / .xlsx files one by one.\nTap Done or type /done when finished."),
    "merge_txt":     ("cmd_merge_txt",     "📚 Merge → TXT\n\nSend .txt / .csv / .vcf / .xlsx files one by one.\nTap Done or type /done when finished."),
    "split_file":    ("cmd_split_file",    "✂️ Split File\n\nSend a .txt / .csv / .vcf / .xlsx file to split."),
    "admin_navy":    ("cmd_admin_navy",    "🛡️ Navy Format\n\nSend a .vcf / .csv / .xlsx file to apply admin navy numbering."),
    "num_compiler":  ("cmd_num_compiler",
                      "📱 Mobile Number Compiler\n\n"
                      "Send a file with mobile numbers:\n"
                      ".txt / .csv / .vcf / .xlsx / .xls\n\n"
                      "Rules:\n"
                      "  • Only mobile numbers allowed\n"
                      "  • Numbers extracted automatically\n\n"
                      "Bot will compile & ask your output format."),
}

# ─────────────────────────────────────────────────────────────
# COMMAND HANDLERS
# ─────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user    = update.effective_user
    user_id = user.id
    context.user_data.pop("state", None)
    context.user_data.pop("awaiting_code", None)

    if is_premium(user_id):
        first = user.first_name or "there"
        await update.message.reply_text(
            "👋 Welcome back, " + first + "!\n"
            "⌨️ Shortcut keyboard is ready." + FOOTER,
            reply_markup=shortcut_keyboard()
        )
        await update.message.reply_text(
            dashboard_text(),
            reply_markup=main_menu_keyboard()
        )
        return

    await update.message.reply_text(
        welcome_text(),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔑 Enter Premium Code", callback_data="enter_code")],
            [InlineKeyboardButton("👤 Contact Admin",      url=ADMIN_URL)],
            [InlineKeyboardButton("ℹ️ Help & Guide",       callback_data="help")],
        ])
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    back_cb = "back_to_welcome"
    if update.callback_query and is_premium(update.callback_query.from_user.id):
        back_cb = "back_to_menu"

    if update.message:
        await update.message.reply_text(
            help_text(),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_menu")]
            ])
        )
    elif update.callback_query:
        await update.callback_query.edit_message_text(
            help_text(),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ Back", callback_data=back_cb)]
            ])
        )

async def reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    msg = (
        "🔄 Session Cleared!\n"
        + THIN + "\n"
        "All temporary data has been wiped.\n"
        "You're ready for a fresh start."
        + FOOTER
    )
    if update.message:
        kb = done_keyboard() if is_premium(update.effective_user.id) else None
        await update.message.reply_text(msg, reply_markup=kb)
    elif update.callback_query:
        kb = done_keyboard() if is_premium(update.callback_query.from_user.id) else None
        await update.callback_query.edit_message_text(msg, reply_markup=kb)

async def done_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_premium(user_id):
        await update.message.reply_text(premium_required_msg())
        return

    state = context.user_data.get("state")
    fcb   = context.user_data.get("active_feature_cb")

    if state == "merge_vcf":
        files = context.user_data.get("merge_vcf_files", [])
        if not files:
            await update.message.reply_text("❌ No VCF files received yet.")
            return
        stems = context.user_data.get("merge_vcf_stems", [])
        context.user_data["pending_result"]     = "\n".join(files)
        context.user_data["pending_stem"]       = "_".join(stems[:3]) if stems else "merged"
        context.user_data["pending_source_ext"] = "vcf"
        context.user_data["state"]              = "ask_output_fmt"
        context.user_data.pop("merge_vcf_files", None)
        context.user_data.pop("merge_vcf_stems", None)
        await update.message.reply_text(
            "✅ " + str(len(files)) + " VCF files queued!\n📤 Choose output format:" + FOOTER,
            reply_markup=output_fmt_keyboard(fcb)
        )

    elif state == "merge_txt":
        files = context.user_data.get("merge_txt_files", [])
        if not files:
            await update.message.reply_text("❌ No TXT files received yet.")
            return
        stems = context.user_data.get("merge_txt_stems", [])
        context.user_data["pending_result"]     = "\n".join(files)
        context.user_data["pending_stem"]       = "_".join(stems[:3]) if stems else "merged"
        context.user_data["pending_source_ext"] = "txt"
        context.user_data["state"]              = "ask_output_fmt"
        context.user_data.pop("merge_txt_files", None)
        context.user_data.pop("merge_txt_stems", None)
        await update.message.reply_text(
            "✅ " + str(len(files)) + " TXT files queued!\n📤 Choose output format:" + FOOTER,
            reply_markup=output_fmt_keyboard(fcb)
        )

    else:
        await update.message.reply_text(
            "⚠️ Nothing to finalize. Start a merge operation first.",
            reply_markup=main_menu_keyboard()
        )

# ─────────────────────────────────────────────────────────────
# FEATURE COMMAND SHORTCUTS
# ─────────────────────────────────────────────────────────────

async def _feature(update: Update, context: ContextTypes.DEFAULT_TYPE, state: str):
    if not is_premium(update.effective_user.id):
        await update.message.reply_text(premium_required_msg())
        return
    fcb, prompt = FEATURE_PROMPTS[state]
    context.user_data["state"]             = state
    context.user_data["active_feature_cb"] = fcb
    await update.message.reply_text(prompt + FOOTER, reply_markup=back_keyboard(fcb))

async def cmd_txt_to_vcf(u, c):     await _feature(u, c, "txt_to_vcf")
async def cmd_txtvcf_to_csv(u, c):  await _feature(u, c, "txtvcf_to_csv")
async def cmd_csv_to_vcf(u, c):     await _feature(u, c, "csv_to_vcf")
async def cmd_vcf_to_txt(u, c):     await _feature(u, c, "vcf_to_txt")
async def cmd_msg_to_txt(u, c):     await _feature(u, c, "msg_to_txt")
async def cmd_rename_file(u, c):    await _feature(u, c, "rename_file")
async def cmd_rename_ctc(u, c):     await _feature(u, c, "rename_ctc")
async def cmd_merge_vcf(u, c):      await _feature(u, c, "merge_vcf")
async def cmd_merge_txt(u, c):      await _feature(u, c, "merge_txt")
async def cmd_split_file(u, c):     await _feature(u, c, "split_file")
async def cmd_admin_navy_file(u, c): await _feature(u, c, "admin_navy")
async def cmd_num_compiler(u, c):   await _feature(u, c, "num_compiler")

# ─────────────────────────────────────────────────────────────
# CALLBACK QUERY HANDLER
# ─────────────────────────────────────────────────────────────

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    data    = query.data
    user_id = query.from_user.id

    if data == "enter_code":
        context.user_data["awaiting_code"] = True
        await query.edit_message_text(
            "🔑 Premium Code Verification\n"
            + THIN + "\n"
            "Please type and send your premium code below 👇"
            + FOOTER,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ Back", callback_data="back_to_welcome")]
            ])
        )
        return

    if data == "help":
        await help_command(update, context)
        return

    if data == "cmd_reset":
        await reset_command(update, context)
        return

    if data == "back_to_menu":
        context.user_data.pop("state", None)
        context.user_data.pop("active_feature_cb", None)
        if is_premium(user_id):
            await query.edit_message_text(
                dashboard_text(),
                reply_markup=main_menu_keyboard()
            )
        else:
            await query.edit_message_text(
                "🔒 Premium Required\n"
                "💎 Get access: " + ADMIN_USERNAME + FOOTER,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔑 Enter Code", callback_data="enter_code")]
                ])
            )
        return

    if data == "back_to_welcome":
        context.user_data.pop("awaiting_code", None)
        context.user_data.pop("state", None)
        await query.edit_message_text(
            welcome_text(),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔑 Enter Premium Code", callback_data="enter_code")],
                [InlineKeyboardButton("👤 Contact Admin",      url=ADMIN_URL)],
                [InlineKeyboardButton("ℹ️ Help & Guide",       callback_data="help")],
            ])
        )
        return

    if data == "merge_vcf_done":
        files = context.user_data.get("merge_vcf_files", [])
        if not files:
            await query.answer("⚠️ No files received yet!", show_alert=True)
            return
        stems = context.user_data.get("merge_vcf_stems", [])
        context.user_data["pending_result"]     = "\n".join(files)
        context.user_data["pending_stem"]       = "_".join(stems[:3]) if stems else "merged"
        context.user_data["pending_source_ext"] = "vcf"
        context.user_data["state"]              = "ask_output_fmt"
        fcb = context.user_data.get("active_feature_cb")
        context.user_data.pop("merge_vcf_files", None)
        context.user_data.pop("merge_vcf_stems", None)
        await query.edit_message_text(
            "✅ " + str(len(files)) + " VCF files ready!\n📤 Choose output format:" + FOOTER,
            reply_markup=output_fmt_keyboard(fcb)
        )
        return

    if data == "merge_txt_done":
        files = context.user_data.get("merge_txt_files", [])
        if not files:
            await query.answer("⚠️ No files received yet!", show_alert=True)
            return
        stems = context.user_data.get("merge_txt_stems", [])
        context.user_data["pending_result"]     = "\n".join(files)
        context.user_data["pending_stem"]       = "_".join(stems[:3]) if stems else "merged"
        context.user_data["pending_source_ext"] = "txt"
        context.user_data["state"]              = "ask_output_fmt"
        fcb = context.user_data.get("active_feature_cb")
        context.user_data.pop("merge_txt_files", None)
        context.user_data.pop("merge_txt_stems", None)
        await query.edit_message_text(
            "✅ " + str(len(files)) + " TXT files ready!\n📤 Choose output format:" + FOOTER,
            reply_markup=output_fmt_keyboard(fcb)
        )
        return

    if data in ("fmt_txt", "fmt_vcf", "fmt_csv", "fmt_xlsx"):
        chosen_ext = data.split("_", 1)[1]
        fcb        = context.user_data.get("active_feature_cb")
        cur_state  = context.user_data.get("state")

        if cur_state == "split_ask_fmt":
            split_data = context.user_data.get("split_file_data")
            split_stem = context.user_data.get("split_file_stem", "split")
            count      = context.user_data.get("split_count", 100)
            orig_ext   = context.user_data.get("split_file_ext", "vcf")
            if not split_data:
                await query.edit_message_text(
                    "❌ File not found. Please start again.",
                    reply_markup=back_keyboard(fcb)
                )
                return
            prog = None
            chunks = split_contacts(split_data, count, orig_ext)
            for i, chunk in enumerate(chunks, 1):
                file_bytes, _ = convert_content(chunk, orig_ext, chosen_ext)
                oname = split_stem + "_part" + str(i) + "." + chosen_ext
                buf = io.BytesIO(file_bytes)
                buf.name = oname
                await query.message.reply_document(document=buf, filename=oname)
            context.user_data.pop("state", None)
            context.user_data.pop("split_file_data", None)
            context.user_data.pop("split_file_ext", None)
            context.user_data.pop("split_file_stem", None)
            context.user_data.pop("split_count", None)
            await query.message.reply_text(
                "🎉 Split Complete!\n"
                + THIN + "\n"
                "📦 Files created: " + str(len(chunks)) + "\n"
                "✂️ Per file: " + str(count) + " contacts\n"
                "📋 Format: ." + chosen_ext + FOOTER,
                reply_markup=done_keyboard(fcb)
            )
            return

        result     = context.user_data.get("pending_result", "")
        stem       = context.user_data.get("pending_stem", "output")
        source_ext = context.user_data.get("pending_source_ext", "vcf")
        input_count = context.user_data.get("pending_input_count", 0)

        prog = None
        file_bytes, _ = convert_content(result, source_ext, chosen_ext)
        oname = stem + "." + chosen_ext

        # Count output records
        if chosen_ext == "vcf":
            out_count = file_bytes.decode("utf-8", errors="ignore").upper().count("BEGIN:VCARD")
        elif chosen_ext == "txt":
            out_count = len([l for l in file_bytes.decode("utf-8", errors="ignore").splitlines() if l.strip()])
        elif chosen_ext in ("csv", "xlsx"):
            out_count = input_count  # same rows, different format
        else:
            out_count = input_count

        buf   = io.BytesIO(file_bytes)
        buf.name = oname
        await query.message.reply_document(
            document=buf,
            filename=oname,
            caption=(
                "✅ File Ready!\n"
                + THIN + "\n"
                "📄 " + oname + "\n"
                "📦 Format: ." + chosen_ext + "\n"
                "📥 Input records: " + str(input_count) + "\n"
                "📤 Output records: " + str(out_count)
                + FOOTER
            )
        )
        context.user_data.pop("state", None)
        context.user_data.pop("pending_result", None)
        context.user_data.pop("pending_stem", None)
        context.user_data.pop("pending_source_ext", None)
        context.user_data.pop("pending_input_count", None)
        context.user_data.pop("fmt_options", None)
        await query.message.reply_text(
            "🎉 All done! What's next?",
            reply_markup=done_keyboard(fcb)
        )
        return

    if data in ("cmp_txt", "cmp_vcf", "cmp_csv", "cmp_xlsx", "cmp_xls"):
        chosen_ext = data.split("_", 1)[1]
        fcb        = context.user_data.get("active_feature_cb")
        numbers    = context.user_data.get("compiler_numbers", [])
        stem       = context.user_data.get("compiler_stem", "compiled")
        tag        = context.user_data.get("compiler_tag", "")
        total      = len(numbers)

        if not numbers:
            await query.edit_message_text(
                "❌ No data found. Please start again.",
                reply_markup=back_keyboard(fcb)
            )
            return

        prog = None

        file_bytes, out_ext = compile_numbers_to_output(numbers, chosen_ext, tag)
        oname = stem + "_compiled." + out_ext
        buf   = io.BytesIO(file_bytes)
        buf.name = oname

        await query.message.reply_document(
            document=buf,
            filename=oname,
            caption=(
                "✅ Number Compiler Done!\n"
                + THIN + "\n"
                "📋 Total Numbers: " + str(total) + "\n"
                "🏷️ Tag: " + (tag if tag else "—") + "\n"
                "📊 Columns: Name | Phone | Country Code | Email | Tags\n"
                "📄 File: " + oname + "\n"
                "📦 Format: ." + out_ext
                + FOOTER
            )
        )
        context.user_data.pop("state", None)
        context.user_data.pop("compiler_numbers", None)
        context.user_data.pop("compiler_stem", None)
        context.user_data.pop("compiler_tag", None)
        await query.message.reply_text(
            "🎉 All done! What's next?",
            reply_markup=done_keyboard(fcb)
        )
        return

    feature_map = {
        "cmd_txt_to_vcf":    "txt_to_vcf",
        "cmd_txtvcf_to_csv": "txtvcf_to_csv",
        "cmd_csv_to_vcf":    "csv_to_vcf",
        "cmd_vcf_to_txt":    "vcf_to_txt",
        "cmd_msg_to_txt":    "msg_to_txt",
        "cmd_rename_file":   "rename_file",
        "cmd_rename_ctc":    "rename_ctc",
        "cmd_merge_vcf":     "merge_vcf",
        "cmd_merge_txt":     "merge_txt",
        "cmd_split_file":    "split_file",
        "cmd_admin_navy":    "admin_navy",
        "cmd_num_compiler":  "num_compiler",
    }

    if data in feature_map:
        if not is_premium(user_id):
            await query.edit_message_text(premium_required_msg())
            return
        state = feature_map[data]
        _, prompt = FEATURE_PROMPTS[state]
        context.user_data["state"]             = state
        context.user_data["active_feature_cb"] = data
        context.user_data.pop("merge_vcf_files", None)
        context.user_data.pop("merge_txt_files", None)
        await query.edit_message_text(
            prompt + FOOTER,
            reply_markup=back_keyboard(data)
        )

# ─────────────────────────────────────────────────────────────
# TEXT MESSAGE HANDLER
# ─────────────────────────────────────────────────────────────

async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text    = update.message.text.strip()
    state   = context.user_data.get("state")

    # Shortcut keyboard button mapping
    shortcut_map = {
        "📁 TXT→VCF":  "txt_to_vcf",
        "📄 VCF→TXT":  "vcf_to_txt",
        "📇 CSV→VCF":  "csv_to_vcf",
        "📊 →CSV":     "txtvcf_to_csv",
        "📝 Rename":   "rename_file",
        "✏️ Ren.CTC":  "rename_ctc",
        "🧩 MergeVCF": "merge_vcf",
        "📚 MergeTXT": "merge_txt",
        "✂️ Split":    "split_file",
        "🛡️ Navy":     "admin_navy",
        "ℹ️ Help":     "__help__",
        "🔄 Reset":    "__reset__",
    }

    if text in shortcut_map:
        mapped = shortcut_map[text]
        if mapped == "__help__":
            await help_command(update, context)
            return
        if mapped == "__reset__":
            await reset_command(update, context)
            return
        if not is_premium(user_id):
            await update.message.reply_text(premium_required_msg())
            return
        fcb, prompt = FEATURE_PROMPTS[mapped]
        context.user_data["state"]             = mapped
        context.user_data["active_feature_cb"] = fcb
        await update.message.reply_text(prompt + FOOTER, reply_markup=back_keyboard(fcb))
        return

    # Premium code verification
    if context.user_data.get("awaiting_code"):
        context.user_data["awaiting_code"] = False
        if text == PREMIUM_CODE:
            premium_users.add(user_id)
            first = update.effective_user.first_name or "there"
            msg = await update.message.reply_text("🔓 Verifying code...")
            await msg.edit_text("✅ Code Accepted!\n🚀 Unlocking premium features...")
            await msg.edit_text("🎉 Premium Access Granted!\n\nWelcome, " + first + "! 👑")
            await update.message.reply_text(
                "⌨️ Shortcut keyboard activated!" + FOOTER,
                reply_markup=shortcut_keyboard()
            )
            await update.message.reply_text(
                dashboard_text(),
                reply_markup=main_menu_keyboard()
            )
        else:
            await update.message.reply_text(
                "❌ Invalid Code\n"
                + THIN + "\n"
                "⚠️ Contact admin: " + ADMIN_USERNAME + FOOTER,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔑 Try Again",     callback_data="enter_code")],
                    [InlineKeyboardButton("👤 Contact Admin", url=ADMIN_URL)],
                ])
            )
        return

    if not is_premium(user_id):
        await update.message.reply_text(
            premium_required_msg(),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔑 Enter Code", callback_data="enter_code")]
            ])
        )
        return

    fcb = context.user_data.get("active_feature_cb")

    if state == "msg_to_txt":
        buf = io.BytesIO(text.encode("utf-8"))
        buf.name = "message.txt"
        await update.message.reply_document(
            document=buf,
            filename="message.txt",
            caption="✅ Message saved as TXT!" + FOOTER
        )
        context.user_data.pop("state", None)
        await update.message.reply_text("🎉 Done! What's next?", reply_markup=done_keyboard(fcb))
        return

    if state == "rename_newname":
        new_name   = text
        file_bytes = context.user_data.get("rename_file_bytes")
        if not file_bytes:
            await update.message.reply_text("❌ File not found. Please start again.", reply_markup=back_keyboard(fcb))
            return
        buf = io.BytesIO(file_bytes)
        buf.name = new_name
        await update.message.reply_document(
            document=buf,
            filename=new_name,
            caption="✅ File renamed to: " + new_name + FOOTER
        )
        context.user_data.pop("state", None)
        context.user_data.pop("rename_file_bytes", None)
        context.user_data.pop("rename_file_stem", None)
        await update.message.reply_text("🎉 Done! What's next?", reply_markup=done_keyboard(fcb))
        return

    if state == "rename_ctc_prefix":
        prefix    = text
        vcf_bytes = context.user_data.get("rename_ctc_bytes")
        stem      = context.user_data.get("rename_ctc_stem", "contacts")
        if not vcf_bytes:
            await update.message.reply_text("❌ VCF not found. Please start again.", reply_markup=back_keyboard(fcb))
            return
        result = rename_vcf_contacts(vcf_bytes.decode("utf-8", errors="ignore"), prefix)
        fname  = stem + "_" + prefix + "_renamed.vcf"
        buf    = io.BytesIO(result.encode("utf-8"))
        buf.name = fname
        await update.message.reply_document(
            document=buf,
            filename=fname,
            caption="✅ Contacts renamed!\n🏷️ Prefix: " + prefix + "\n📄 " + fname + FOOTER
        )
        context.user_data.pop("state", None)
        context.user_data.pop("rename_ctc_bytes", None)
        context.user_data.pop("rename_ctc_stem", None)
        await update.message.reply_text("🎉 Done! What's next?", reply_markup=done_keyboard(fcb))
        return

    if state == "split_count":
        try:
            count = int(text)
            if count <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text(
                "❌ Please enter a valid positive number.\nExample: 100",
                reply_markup=back_keyboard(fcb)
            )
            return
        split_data = context.user_data.get("split_file_data")
        if not split_data:
            await update.message.reply_text("❌ File not found. Please start again.", reply_markup=back_keyboard(fcb))
            return
        context.user_data["split_count"] = count
        context.user_data["state"]       = "split_ask_fmt"
        total = count_contacts(split_data, context.user_data.get("split_file_ext", "txt"))
        files_count = (total + count - 1) // count if total > 0 else "?"
        await update.message.reply_text(
            "✅ Got it!\n"
            + THIN + "\n"
            "📋 Total Contacts: " + str(total) + "\n"
            "✂️ Per File: " + str(count) + "\n"
            "📦 Files to create: " + str(files_count) + "\n\n"
            "📤 Choose output format:" + FOOTER,
            reply_markup=output_fmt_keyboard(fcb)
        )
        return

    if state == "compiler_ask_tag":
        tag     = text.strip()
        numbers = context.user_data.get("compiler_numbers", [])
        fcb     = context.user_data.get("active_feature_cb")
        if not numbers:
            await update.message.reply_text("❌ No data found. Please start again.", reply_markup=back_keyboard(fcb))
            return
        context.user_data["compiler_tag"] = tag
        context.user_data["state"]        = "compiler_ask_fmt"
        total = len(numbers)
        await update.message.reply_text(
            "✅ Tag set: " + tag + "\n"
            + THIN + "\n"
            "📋 Total Numbers: " + str(total) + "\n\n"
            "📤 Choose output format:" + FOOTER,
            reply_markup=compiler_fmt_keyboard(fcb)
        )
        return

    await update.message.reply_text(
        "💡 Select a feature from the menu to get started.",
        reply_markup=main_menu_keyboard()
    )

# ─────────────────────────────────────────────────────────────
# FILE / DOCUMENT HANDLER
# ─────────────────────────────────────────────────────────────

async def file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if not is_premium(user_id):
        await update.message.reply_text(
            premium_required_msg(),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔑 Enter Code", callback_data="enter_code")]
            ])
        )
        return

    state = context.user_data.get("state")
    doc   = update.message.document

    if not doc:
        await update.message.reply_text("❌ Please send a valid file.")
        return

    fname = doc.file_name or "file"
    ext   = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
    fcb   = context.user_data.get("active_feature_cb")

    prog = None

    try:
        tg_file = await doc.get_file()
        raw     = await tg_file.download_as_bytearray()
        raw     = bytes(raw)

        # ── Validate extension is one of the 4 accepted formats ──
        if ext not in ALL_FORMATS:
            await update.message.reply_text("↩️ Try again:", reply_markup=back_keyboard(fcb))
            return

        # ── Read & normalise the file ─────────────────────────
        content, norm_ext = read_file_content(raw, ext)

        # ── Route by active state ─────────────────────────────

        if state == "txt_to_vcf":
            total_in = count_contacts(content, norm_ext)
            context.user_data["pending_result"]     = content
            context.user_data["pending_stem"]       = file_stem(fname)
            context.user_data["pending_source_ext"] = norm_ext
            context.user_data["pending_input_count"] = total_in
            context.user_data["state"]              = "ask_output_fmt"
            await update.message.reply_text(
                "📊 Input Analysis\n"
                + THIN + "\n"
                "📋 Records found: " + str(total_in) + "\n\n"
                "📤 Choose output format:" + FOOTER,
                reply_markup=output_fmt_keyboard(fcb)
            )

        elif state == "txtvcf_to_csv":
            total_in = count_contacts(content, norm_ext)
            context.user_data["pending_result"]     = content
            context.user_data["pending_stem"]       = file_stem(fname)
            context.user_data["pending_source_ext"] = norm_ext
            context.user_data["pending_input_count"] = total_in
            context.user_data["state"]              = "ask_output_fmt"
            await update.message.reply_text(
                "📊 Input Analysis\n"
                + THIN + "\n"
                "📋 Records found: " + str(total_in) + "\n\n"
                "📤 Choose output format:" + FOOTER,
                reply_markup=output_fmt_keyboard(fcb)
            )

        elif state == "csv_to_vcf":
            total_in = count_contacts(content, norm_ext)
            context.user_data["pending_result"]     = content
            context.user_data["pending_stem"]       = file_stem(fname)
            context.user_data["pending_source_ext"] = norm_ext
            context.user_data["pending_input_count"] = total_in
            context.user_data["state"]              = "ask_output_fmt"
            await update.message.reply_text(
                "📊 Input Analysis\n"
                + THIN + "\n"
                "📋 Records found: " + str(total_in) + "\n\n"
                "📤 Choose output format:" + FOOTER,
                reply_markup=output_fmt_keyboard(fcb)
            )

        elif state == "vcf_to_txt":
            total_in = count_contacts(content, norm_ext)
            context.user_data["pending_result"]     = content
            context.user_data["pending_stem"]       = file_stem(fname)
            context.user_data["pending_source_ext"] = norm_ext
            context.user_data["pending_input_count"] = total_in
            context.user_data["state"]              = "ask_output_fmt"
            await update.message.reply_text(
                "📊 Input Analysis\n"
                + THIN + "\n"
                "📋 Records found: " + str(total_in) + "\n\n"
                "📤 Choose output format:" + FOOTER,
                reply_markup=output_fmt_keyboard(fcb)
            )

        elif state == "rename_file":
            context.user_data["rename_file_bytes"] = raw
            context.user_data["rename_file_stem"]  = file_stem(fname)
            context.user_data["state"] = "rename_newname"
            await update.message.reply_text(
                "📝 Got: " + fname + "\n\n"
                "Now send the new filename with extension.\n"
                "Example: contacts_backup.vcf" + FOOTER,
                reply_markup=back_keyboard(fcb)
            )

        elif state == "rename_ctc":
            # For rename contacts, normalise to VCF text first
            if norm_ext == "vcf":
                vcf_text = content
            elif norm_ext in ("csv", "xlsx"):
                vcf_text = csv_to_vcf(content)
            else:
                vcf_text = txt_to_vcf(content)
            total_in = vcf_text.upper().count("BEGIN:VCARD")
            context.user_data["rename_ctc_bytes"] = vcf_text.encode("utf-8")
            context.user_data["rename_ctc_stem"]  = file_stem(fname)
            context.user_data["state"] = "rename_ctc_prefix"
            await update.message.reply_text(
                "✏️ Got: " + fname + "\n"
                "📋 Contacts found: " + str(total_in) + "\n\n"
                "Now send the contact name prefix.\n"
                "Example: GGOD → GGOD 1, GGOD 2..." + FOOTER,
                reply_markup=back_keyboard(fcb)
            )

        elif state == "merge_vcf":
            files = context.user_data.setdefault("merge_vcf_files", [])
            stems = context.user_data.setdefault("merge_vcf_stems", [])
            files.append(content)
            stems.append(file_stem(fname))
            this_count = count_contacts(content, norm_ext)
            running_total = sum(count_contacts(f, norm_ext) for f in files)
            await update.message.reply_text(
                "📎 " + fname + " added ✓\n"
                "📋 This file: " + str(this_count) + " contacts\n"
                "📦 Files: " + str(len(files)) + "  |  Total contacts: " + str(running_total) + "\n\n"
                "Send more files or tap Done to merge." + FOOTER,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ Done — Merge Now", callback_data="merge_vcf_done")],
                    [InlineKeyboardButton("◀️ Cancel",           callback_data="back_to_menu")],
                ])
            )

        elif state == "merge_txt":
            files = context.user_data.setdefault("merge_txt_files", [])
            stems = context.user_data.setdefault("merge_txt_stems", [])
            files.append(content)
            stems.append(file_stem(fname))
            this_count = count_contacts(content, norm_ext)
            running_total = sum(count_contacts(f, norm_ext) for f in files)
            await update.message.reply_text(
                "📎 " + fname + " added ✓\n"
                "📋 This file: " + str(this_count) + " records\n"
                "📦 Files: " + str(len(files)) + "  |  Total records: " + str(running_total) + "\n\n"
                "Send more files or tap Done to merge." + FOOTER,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ Done — Merge Now", callback_data="merge_txt_done")],
                    [InlineKeyboardButton("◀️ Cancel",           callback_data="back_to_menu")],
                ])
            )

        elif state == "split_file":
            context.user_data["split_file_data"] = content
            context.user_data["split_file_ext"]  = norm_ext
            context.user_data["split_file_stem"] = file_stem(fname)
            context.user_data["state"] = "split_count"
            total = count_contacts(content, norm_ext)
            await update.message.reply_text(
                "✂️ Got: " + fname + "\n\n"
                "📊 Analysis Complete\n"
                + THIN + "\n"
                "📋 Total Contacts: " + str(total) + "\n\n"
                "🔢 Enter how many contacts per file?\n"
                "Example: 100" + FOOTER,
                reply_markup=back_keyboard(fcb)
            )

        elif state == "admin_navy":
            # Normalise to VCF entries first
            if norm_ext == "vcf":
                vcf_text = content
            elif norm_ext in ("csv", "xlsx"):
                vcf_text = csv_to_vcf(content)
            else:
                vcf_text = txt_to_vcf(content)
            total_in = vcf_text.upper().count("BEGIN:VCARD")
            navy_result = admin_navy_format(vcf_text)
            context.user_data["pending_result"]      = navy_result
            context.user_data["pending_stem"]        = file_stem(fname)
            context.user_data["pending_source_ext"]  = "vcf"
            context.user_data["pending_input_count"] = total_in
            context.user_data["state"]               = "ask_output_fmt"
            await update.message.reply_text(
                "📊 Input Analysis\n"
                + THIN + "\n"
                "📋 Contacts processed: " + str(total_in) + "\n\n"
                "📤 Choose output format:" + FOOTER,
                reply_markup=output_fmt_keyboard(fcb)
            )

        elif state == "num_compiler":
            # Accept all 5 formats
            if ext not in COMPILER_FMTS:
                await update.message.reply_text("↩️ Try again:", reply_markup=back_keyboard(fcb))
                return

            # Extract numbers from any format
            if norm_ext == "vcf":
                numbers_raw = [l.split(":")[-1].strip()
                               for l in content.splitlines()
                               if l.strip().upper().startswith("TEL")]
            elif norm_ext == "csv":
                rows = list(csv.reader(io.StringIO(content)))
                numbers_raw = []
                if rows:
                    # Try to find phone column
                    header = [h.strip().lower() for h in rows[0]]
                    phone_idx = None
                    for candidate in ("mobile number", "phone", "tel", "number", "mobile"):
                        if candidate in header:
                            phone_idx = header.index(candidate)
                            break
                    for row in rows[1:]:
                        if phone_idx is not None and phone_idx < len(row):
                            val = row[phone_idx].strip()
                        elif row:
                            # fallback: take first non-empty cell
                            val = next((c.strip() for c in row if c.strip()), "")
                        else:
                            val = ""
                        if val:
                            numbers_raw.append(val)
            else:
                # txt
                numbers_raw = [l.strip() for l in content.splitlines() if l.strip()]

            # Validate
            invalid = [n for n in numbers_raw if not is_valid_number(n)]
            if invalid:
                sample = "\n".join(invalid[:5])
                await update.message.reply_text(
                    "Fix the file and send again.",
                    reply_markup=back_keyboard(fcb)
                )
                return

            numbers = [normalize_number(n) for n in numbers_raw]
            total   = len(numbers)

            # Store numbers and ask for tag first
            context.user_data["compiler_numbers"] = numbers
            context.user_data["compiler_stem"]    = file_stem(fname)
            context.user_data["state"]             = "compiler_ask_tag"

            await update.message.reply_text(
                "📱 Numbers Loaded!\n"
                + THIN + "\n"
                "📋 Total Numbers: " + str(total) + "\n\n"
                "🏷️ Enter a Tag for all contacts:\n"
                "(This tag will be applied to every row)\n"
                "Example: VIP, Client, India2025" + FOOTER,
                reply_markup=back_keyboard(fcb)
            )

        else:
            await update.message.reply_text("👇 Select a feature:", reply_markup=main_menu_keyboard())

    except Exception as e:
        logger.error("File handler error: %s", e)
        try:
            pass
        except Exception:
            pass
        await update.message.reply_text(
            "❌ Something went wrong. Please try again.",
            reply_markup=back_keyboard(fcb)
        )

# ─────────────────────────────────────────────────────────────
# ERROR HANDLER
# ─────────────────────────────────────────────────────────────

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error("Unhandled exception: %s", context.error, exc_info=context.error)

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    print("=" * 40)
    print("  ⚡ GGOD FATHERR BOT STARTING...  ")
    print("  👑 PREMIUM SYSTEM ACTIVE         ")
    print("  🟢 STATUS: ONLINE                ")
    print("=" * 40)

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start",           start))
    app.add_handler(CommandHandler("help",            help_command))
    app.add_handler(CommandHandler("reset",           reset_command))
    app.add_handler(CommandHandler("done",            done_command))
    app.add_handler(CommandHandler("txt_to_vcf",      cmd_txt_to_vcf))
    app.add_handler(CommandHandler("txtvcf_to_csv",   cmd_txtvcf_to_csv))
    app.add_handler(CommandHandler("csv_to_vcf",      cmd_csv_to_vcf))
    app.add_handler(CommandHandler("vcf_to_txt",      cmd_vcf_to_txt))
    app.add_handler(CommandHandler("msg_to_txt",      cmd_msg_to_txt))
    app.add_handler(CommandHandler("rename_file",     cmd_rename_file))
    app.add_handler(CommandHandler("rename_ctc",      cmd_rename_ctc))
    app.add_handler(CommandHandler("merge_vcf",       cmd_merge_vcf))
    app.add_handler(CommandHandler("merge_txt",       cmd_merge_txt))
    app.add_handler(CommandHandler("split_file",      cmd_split_file))
    app.add_handler(CommandHandler("admin_navy_file", cmd_admin_navy_file))
    app.add_handler(CommandHandler("num_compiler",    cmd_num_compiler))

    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, file_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    app.add_error_handler(error_handler)

    print("🤖 Bot is polling for updates...")
    app.run_polling(allowed_updates=["message", "callback_query"])

if __name__ == "__main__":
    main()
